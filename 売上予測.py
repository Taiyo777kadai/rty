import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import LabelEncoder
import math

excel_path = r'売り上げサンプル.xlsx'
sheet_name = '売上データ'

st.title('売上予測アプリ')

youbi = st.selectbox('曜日区分', ['平日', '休日', '祝祭日'])
tenki = st.selectbox('天気', ['晴れ', '雨', '曇り'])

if 'yoso_100' not in st.session_state:
    st.session_state['yoso_100'] = None
if 'shiire_list' not in st.session_state:
    st.session_state['shiire_list'] = None
if 'name_list' not in st.session_state:
    st.session_state['name_list'] = None
if 'tanka_list' not in st.session_state:
    st.session_state['tanka_list'] = None
if 'prev_yoso' not in st.session_state:
    st.session_state['prev_yoso'] = 0
if 'prev_jissai' not in st.session_state:
    st.session_state['prev_jissai'] = 0

if st.button('予測'):
    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        le_youbi = LabelEncoder()
        le_tenki = LabelEncoder()
        df['曜日区分'] = le_youbi.fit_transform(df['曜日区分'])
        df['天気'] = le_tenki.fit_transform(df['天気'])
        if '前回予測' not in df.columns:
            df['前回予測'] = 0
        if '前回実績' not in df.columns:
            df['前回実績'] = 0
        X = df[['曜日区分', '天気', '前回予測', '前回実績']]
        y = df['売上']
        model = LinearRegression()
        model.fit(X, y)
        input_youbi = le_youbi.transform([youbi])
        input_tenki = le_tenki.transform([tenki])
        if len(df) > 0:
            prev_yoso = df.iloc[-1]['前回予測']
            prev_jissai = df.iloc[-1]['前回実績']
        else:
            prev_yoso = 0
            prev_jissai = 0
        yoso = model.predict([[input_youbi[0], input_tenki[0], prev_yoso, prev_jissai]])[0]
        yoso_100 = int(round(yoso / 100) * 100)
        st.success(f'{youbi}・{tenki}の予測売上は {yoso_100} 円です。')

        wb = load_workbook(excel_path)
        menu_ws = wb['メニュー']
        menu_header = [cell.value for cell in menu_ws[1]]
        tanka_col = menu_header.index('単価') + 1
        name_col = menu_header.index('メニュー名') + 1

        menu_rows = []
        tanka_list = []
        name_list = []
        for row in range(2, menu_ws.max_row + 1):
            tanka = menu_ws.cell(row=row, column=tanka_col).value
            name = menu_ws.cell(row=row, column=name_col).value
            if tanka is not None and tanka != 0:
                tanka_list.append(tanka)
                name_list.append(name)
                menu_rows.append(row)
            else:
                tanka_list.append(0)
                name_list.append(name)
                menu_rows.append(row)

        menu_count = len([t for t in tanka_list if t > 0])
        if menu_count == 0:
            shiire_list = [0 for _ in tanka_list]
        else:
            budget_per_menu = yoso_100 // menu_count
            shiire_list = [budget_per_menu // tanka if tanka > 0 else 0 for tanka in tanka_list]
            total = sum([shiire * tanka for shiire, tanka in zip(shiire_list, tanka_list)])
            nokori = yoso_100 - total
            if tanka_list[-1] > 0:
                add = nokori // tanka_list[-1]
                shiire_list[-1] += add
                total += add * tanka_list[-1]

        result_df = pd.DataFrame({
            'メニュー名': name_list,
            '単価': tanka_list,
            '仕入れ数': shiire_list
        })
        st.subheader('予測仕入れ数')
        st.table(result_df)

        st.session_state['yoso_100'] = yoso_100
        st.session_state['shiire_list'] = shiire_list
        st.session_state['name_list'] = name_list
        st.session_state['tanka_list'] = tanka_list
        st.session_state['prev_yoso'] = prev_yoso
        st.session_state['prev_jissai'] = prev_jissai

    except Exception as e:
        st.error(f'エラー: {e}')

if st.session_state['yoso_100'] is not None:
    jissai_uriage = st.number_input('実際の売上（数字）', min_value=0, step=100, key='jissai_input')
    if st.button('記録'):
        try:
            wb = load_workbook(excel_path)
            ws = wb[sheet_name]
            ws.append([
                youbi,
                tenki,
                st.session_state['yoso_100'],
                jissai_uriage,
                st.session_state['prev_yoso'],
                st.session_state['prev_jissai']
            ])
            menu_ws = wb['メニュー']
            menu_header = [cell.value for cell in menu_ws[1]]
            shiire_col = len(menu_header) + 1
            if '仕入れ数' not in menu_header:
                menu_ws.cell(row=1, column=shiire_col, value='仕入れ数')
            for row, shiire in zip(range(2, menu_ws.max_row + 1), st.session_state['shiire_list']):
                menu_ws.cell(row=row, column=shiire_col, value=shiire)
            wb.save(excel_path)
            st.success('予測値・実際の売上・仕入れ数をExcelに記載しました。')
            st.session_state['yoso_100'] = None
        except Exception as e:
            st.error(f'エラー: {e}')